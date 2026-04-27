from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from .models import RFQEntry, Supplier, SupplierContact
from .forms import RFQEntryForm

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import datetime
import json
import re

# Maps Excel column headers → model field names
BULK_COLUMN_MAP = {
    'Supplier Code':                    'supplier_code',
    'Supplier Name':                    'supplier_name',
    'Part No':                          'part_no',
    'Part Description':                 'part_description',
    'Order Qty':                        'order_qty',
    'UOM':                              'uom',
    'Unit Price':                       'unit_price',
    'Currency':                         'currency',
    'PIC':                              'pic',
    'Contact Email':                    'contact_email',
    'Contact Secondary Email':          'contact_secondary_email',
    'Lead Time (days)':                 'lead_time_days',
    'Ship Lead Time (days)':            'ship_lead_time_days',
    'UOM (Quote)':                      'quote_uom',
    'COO':                              'coo',
    'Currency (Quote)':                 'quote_currency',
    'Unit Price 1':                     'unit_price_1',
    'MOQ-1':                            'moq_1',
    'Unit Price 2':                     'unit_price_2',
    'MOQ-2':                            'moq_2',
    'Unit Price 3':                     'unit_price_3',
    'MOQ-3':                            'moq_3',
    'Lot Size':                         'lot_size',
    'HTS Code':                         'hts_code',
    'ECCN/EAR99':                       'eccn_ear99',
    'Manufacture Part Number':          'manufacture_part_number',
    'Manufacturer Name':                'manufacturer_name',
    'Manufacturer Address':             'manufacturer_address',
    'Item Weight (kg)':                 'item_weight_kg',
    'Volume Weight (kg)':               'volume_weight_kg',
    'Russian Steel Confirmation':       'russian_steel_confirmation',
    'Hazmat (Y/N)':                     'hazmat',
    'UN# · SDS/MSDS':                   'un_sds_msds',
    'Product Regulation':               'product_regulation',
    'EOL Status':                       'eol_status',
    'Alternative Part(s)':              'alternative_parts',
    'Alternative Part No':              'alternative_part_no',
    'Mfg Address + Postal (CN Only)':   'mfg_address_postal_cn',
    'UFLPA Compliance (CN Only)':       'uflpa_compliance',
    'UFLPA Start Date':                 'uflpa_start_date',
    'UFLPA Expiry Date':                'uflpa_expiry_date',
    'USMCA Certificate (CA/MX Only)':   'usmca_certificate',
    'USMCA Start Date':                 'usmca_start_date',
    'USMCA Expiry Date':                'usmca_expiry_date',
    'Status':                           'status',
    'Comments':                         'comments',
}

# Known alternative spellings / truncations in supplier Excel templates
HEADER_ALIASES = {
    'Contact Secondar Email':   'Contact Secondary Email',
    'Contact Secondar email':   'Contact Secondary Email',
    'Lead Time':                'Lead Time (days)',
    'Ship Lead Time':           'Ship Lead Time (days)',
    'MOQ -1':                   'MOQ-1',
    'MOQ -2':                   'MOQ-2',
    'MOQ -3':                   'MOQ-3',
    'MOQ - 1':                  'MOQ-1',
    'MOQ - 2':                  'MOQ-2',
    'MOQ - 3':                  'MOQ-3',
    # Merged two-row headers (row1 + row2 combined text)
    'Manufacturer Address (Street|City|ZIP|Country)': 'Manufacturer Address',
    'Start Date':                   'UFLPA Start Date',
    'Expiry Date':                  'UFLPA Expiry Date',
    'Start Date (MM/DD/YYYY)':      'UFLPA Start Date',
    'Expiry Date (MM/DD/YYYY)':     'UFLPA Expiry Date',
    # Wrapped-cell variants — Excel newline inside "(MM/DD/YYY\nY)" → space after normalization
    'Start Date (MM/DD/YYY Y)':     'UFLPA Start Date',
    'Expiry Date (MM/DD/YYY Y)':    'UFLPA Expiry Date',
    'Start Date (MM/DD/YY YY)':     'UFLPA Start Date',
    'Expiry Date (MM/DD/YY YY)':    'UFLPA Expiry Date',
    # Columns that may appear without their parenthetical suffix
    'UFLPA Compliance':                     'UFLPA Compliance (CN Only)',
    'UFLPA Compliance Statement':           'UFLPA Compliance (CN Only)',
    'UFLPA Compliance Statement (CN Only)': 'UFLPA Compliance (CN Only)',
    'Mfg Address + Postal':                 'Mfg Address + Postal (CN Only)',
    'USMCA Certificate':                    'USMCA Certificate (CA/MX Only)',
}

# When a header resolves to a field that was already assigned,
# map it to its second-occurrence counterpart (column-order dependent).
DUPLICATE_FIELD_MAP = {
    'uom':                'quote_uom',
    'currency':           'quote_currency',
    'uflpa_start_date':   'usmca_start_date',
    'uflpa_expiry_date':  'usmca_expiry_date',
}

DECIMAL_FIELDS = {
    'order_qty', 'unit_price', 'unit_price_1', 'moq_1',
    'unit_price_2', 'moq_2', 'unit_price_3', 'moq_3',
    'lot_size', 'item_weight_kg', 'volume_weight_kg',
}
INT_FIELDS = {'lead_time_days', 'ship_lead_time_days'}
DATE_FIELDS = {'uflpa_start_date', 'uflpa_expiry_date', 'usmca_start_date', 'usmca_expiry_date'}


def _normalize_header(text):
    """Collapse newlines / extra whitespace into a single space, strip
    dropdown-arrow indicator characters (▼ ▾ ▽) that Excel embeds."""
    if not text:
        return ''
    t = str(text)
    # Strip common dropdown indicator characters
    for ch in '\u25bc\u25be\u25bd\u25b6\u2193':   # ▼ ▾ ▽ ▶ ↓
        t = t.replace(ch, '')
    return ' '.join(t.split()).strip()


def _resolve_header(text):
    """Return the model field name for a header string, or None."""
    norm = _normalize_header(text)
    if norm in BULK_COLUMN_MAP:
        return BULK_COLUMN_MAP[norm]
    # Check aliases (maps alt text → canonical key)
    canonical = HEADER_ALIASES.get(norm)
    if canonical and canonical in BULK_COLUMN_MAP:
        return BULK_COLUMN_MAP[canonical]
    # Fallback: normalise spaces around hyphens (e.g. "MOQ -1" → "MOQ-1")
    collapsed = re.sub(r'\s*-\s*', '-', norm)
    if collapsed in BULK_COLUMN_MAP:
        return BULK_COLUMN_MAP[collapsed]
    canonical2 = HEADER_ALIASES.get(collapsed)
    if canonical2 and canonical2 in BULK_COLUMN_MAP:
        return BULK_COLUMN_MAP[canonical2]
    # Last resort: strip the trailing parenthetical entirely.
    # Handles wrapped cells where "(MM/DD/YYYY)" becomes "(MM/DD/YYY Y)" or similar.
    stripped = re.sub(r'\s*\(.*', '', norm).strip()
    if stripped and stripped != norm:
        if stripped in BULK_COLUMN_MAP:
            return BULK_COLUMN_MAP[stripped]
        canonical3 = HEADER_ALIASES.get(stripped)
        if canonical3 and canonical3 in BULK_COLUMN_MAP:
            return BULK_COLUMN_MAP[canonical3]
    return None


def _coerce(field, value):
    if value is None or str(value).strip() == '':
        # Numeric/date fields accept NULL; CharFields must use empty string
        if field in DECIMAL_FIELDS or field in INT_FIELDS or field in DATE_FIELDS:
            return None
        return ''
    if field in DECIMAL_FIELDS:
        try:
            return float(str(value).replace(',', ''))
        except (ValueError, TypeError):
            return None
    if field in INT_FIELDS:
        try:
            return int(float(str(value).replace(',', '')))
        except (ValueError, TypeError):
            return None
    if field in DATE_FIELDS:
        if isinstance(value, (datetime.date, datetime.datetime)):
            return value if isinstance(value, datetime.date) else value.date()
        s = str(value).strip()
        if not s: return None
        # Strip time component if present (e.g. "2021-01-29T00:00:00" → "2021-01-29")
        s = s.split('T')[0].strip()
        # Column headers say MM/DD/YYYY — try that first
        for fmt in ('%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d', '%m/%d/%y', '%d/%m/%y'):
            try:
                return datetime.datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None
    return str(value).strip()


def _entry_to_dict(entry):
    def fd(d): return d.isoformat() if d else ''
    def fn(v): return str(v) if v is not None else ''
    return {
        'pk': entry.pk,
        'supplier_code': entry.supplier_code or '',
        'supplier_name': entry.supplier_name or '',
        'part_no': entry.part_no or '',
        'part_description': entry.part_description or '',
        'order_qty': fn(entry.order_qty),
        'uom': entry.uom or '',
        'unit_price': fn(entry.unit_price),
        'currency': entry.currency or '',
        'pic': entry.pic or '',
        'contact_email': entry.contact_email or '',
        'contact_secondary_email': entry.contact_secondary_email or '',
        'lead_time_days': fn(entry.lead_time_days),
        'ship_lead_time_days': fn(entry.ship_lead_time_days),
        'quote_uom': entry.quote_uom or '',
        'coo': entry.coo or '',
        'quote_currency': entry.quote_currency or '',
        'unit_price_1': fn(entry.unit_price_1),
        'moq_1': fn(entry.moq_1),
        'unit_price_2': fn(entry.unit_price_2),
        'moq_2': fn(entry.moq_2),
        'unit_price_3': fn(entry.unit_price_3),
        'moq_3': fn(entry.moq_3),
        'lot_size': fn(entry.lot_size),
        'hts_code': entry.hts_code or '',
        'eccn_ear99': entry.eccn_ear99 or '',
        'manufacture_part_number': entry.manufacture_part_number or '',
        'manufacturer_name': entry.manufacturer_name or '',
        'manufacturer_address': entry.manufacturer_address or '',
        'item_weight_kg': fn(entry.item_weight_kg),
        'volume_weight_kg': fn(entry.volume_weight_kg),
        'russian_steel_confirmation': entry.russian_steel_confirmation or '',
        'hazmat': entry.hazmat or '',
        'un_sds_msds': entry.un_sds_msds or '',
        'product_regulation': entry.product_regulation or '',
        'eol_status': entry.eol_status or '',
        'alternative_parts': entry.alternative_parts or '',
        'alternative_part_no': entry.alternative_part_no or '',
        'mfg_address_postal_cn': entry.mfg_address_postal_cn or '',
        'uflpa_compliance': entry.uflpa_compliance or '',
        'uflpa_start_date': fd(entry.uflpa_start_date),
        'uflpa_expiry_date': fd(entry.uflpa_expiry_date),
        'usmca_certificate': entry.usmca_certificate or '',
        'usmca_start_date': fd(entry.usmca_start_date),
        'usmca_expiry_date': fd(entry.usmca_expiry_date),
        'status': entry.status or '',
        'comments': entry.comments or '',
    }


@login_required
def rfq_list(request):
    form = RFQEntryForm()
    edit_form = RFQEntryForm(prefix='edit')
    total_count = RFQEntry.objects.count()
    KNOWN = {'Completed', 'Partially Data Received', 'No Response Yet'}
    ctx = {
        'total_count': total_count,
        'form': form,
        'edit_form': edit_form,
        'stats': {
            'total_suppliers': RFQEntry.objects.values('supplier_code').distinct().count(),
            'sent':      RFQEntry.objects.exclude(status='').count(),
            'not_sent':  RFQEntry.objects.filter(status='').count(),
            'partial':   RFQEntry.objects.filter(status='Partially Data Received').count(),
            'completed': RFQEntry.objects.filter(status='Completed').count(),
            'no_resp':   RFQEntry.objects.filter(status='No Response Yet').count(),
            'not_valid': RFQEntry.objects.exclude(status='').exclude(status__in=KNOWN).count(),
        },
    }
    return render(request, 'tracker/rfq_list.html', ctx)


@login_required
def rfq_stats(request):
    KNOWN = {'Completed', 'Partially Data Received', 'No Response Yet'}
    return JsonResponse({
        'total_suppliers': RFQEntry.objects.values('supplier_code').distinct().count(),
        'sent':      RFQEntry.objects.exclude(status='').count(),
        'not_sent':  RFQEntry.objects.filter(status='').count(),
        'partial':   RFQEntry.objects.filter(status='Partially Data Received').count(),
        'completed': RFQEntry.objects.filter(status='Completed').count(),
        'no_resp':   RFQEntry.objects.filter(status='No Response Yet').count(),
        'not_valid': RFQEntry.objects.exclude(status='').exclude(status__in=KNOWN).count(),
    })


@login_required
def rfq_data(request):
    from django.db.models import Q
    from django.core.paginator import Paginator

    q = request.GET.get('q', '').strip()
    try:
        page = int(request.GET.get('page', 1))
    except (ValueError, TypeError):
        page = 1
    page_size = request.GET.get('page_size', '20')
    try:
        sort_col = int(request.GET.get('sort_col', -1))
    except (ValueError, TypeError):
        sort_col = -1
    sort_dir = request.GET.get('sort_dir', 'asc')

    qs = RFQEntry.objects.all()

    if q:
        qs = qs.filter(
            Q(supplier_code__icontains=q) | Q(supplier_name__icontains=q) |
            Q(part_no__icontains=q) | Q(part_description__icontains=q) |
            Q(manufacture_part_number__icontains=q) | Q(manufacturer_name__icontains=q) |
            Q(pic__icontains=q) | Q(coo__icontains=q) | Q(hts_code__icontains=q) |
            Q(eccn_ear99__icontains=q) | Q(status__icontains=q) | Q(comments__icontains=q)
        )

    COL_FIELD = {
        0: 'pk', 1: 'supplier_code', 2: 'supplier_name', 3: 'part_no',
        4: 'part_description', 5: 'order_qty', 6: 'uom', 7: 'unit_price',
        8: 'currency', 9: 'pic', 10: 'contact_email', 11: 'contact_secondary_email',
        12: 'lead_time_days', 13: 'ship_lead_time_days', 14: 'quote_uom',
        15: 'coo', 16: 'quote_currency', 17: 'unit_price_1', 18: 'moq_1',
        19: 'unit_price_2', 20: 'moq_2', 21: 'unit_price_3', 22: 'moq_3',
        23: 'lot_size', 24: 'hts_code', 25: 'eccn_ear99',
        26: 'manufacture_part_number', 27: 'manufacturer_name',
        28: 'manufacturer_address', 29: 'item_weight_kg', 30: 'volume_weight_kg',
        31: 'russian_steel_confirmation', 32: 'hazmat', 33: 'un_sds_msds',
        34: 'product_regulation', 35: 'eol_status', 36: 'alternative_parts',
        37: 'alternative_part_no', 38: 'mfg_address_postal_cn',
        39: 'uflpa_compliance', 40: 'uflpa_start_date', 41: 'uflpa_expiry_date',
        42: 'usmca_certificate', 43: 'usmca_start_date', 44: 'usmca_expiry_date',
        45: 'status', 46: 'comments',
    }

    field = COL_FIELD.get(sort_col)
    if field:
        order = field if sort_dir == 'asc' else '-' + field
        qs = qs.order_by(order)

    total = qs.count()

    if page_size == 'all':
        entries = list(qs)
        pages = 1
        page = 1
    else:
        try:
            page_size_int = int(page_size)
        except (ValueError, TypeError):
            page_size_int = 20
        paginator = Paginator(qs, page_size_int)
        page_obj = paginator.get_page(page)
        entries = list(page_obj.object_list)
        pages = paginator.num_pages
        page = page_obj.number

    return JsonResponse({
        'entries': [_entry_to_dict(e) for e in entries],
        'total': total,
        'page': page,
        'pages': pages,
    })


@login_required
def rfq_add(request):
    if request.method == 'POST':
        form = RFQEntryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'RFQ entry added successfully.')
            return redirect('rfq_list')
        else:
            messages.error(request, 'Please correct the errors below.')
            total_count = RFQEntry.objects.count()
            edit_form = RFQEntryForm(prefix='edit')
            return render(request, 'tracker/rfq_list.html', {'total_count': total_count, 'form': form, 'edit_form': edit_form})
    return redirect('rfq_list')


@login_required
def rfq_delete(request, pk):
    entry = get_object_or_404(RFQEntry, pk=pk)
    if request.method == 'POST':
        entry.delete()
        messages.success(request, f'Entry for "{entry.supplier_name} — {entry.part_no}" deleted.')
    return redirect('rfq_list')


@login_required
def rfq_edit(request, pk):
    entry = get_object_or_404(RFQEntry, pk=pk)
    if request.method == 'POST':
        form = RFQEntryForm(request.POST, instance=entry)
        if form.is_valid():
            form.save()
            messages.success(request, 'Entry updated successfully.')
            return redirect('rfq_list')
    else:
        form = RFQEntryForm(instance=entry)
    return render(request, 'tracker/rfq_edit.html', {'form': form, 'entry': entry})


@login_required
def rfq_export(request):
    entries   = RFQEntry.objects.all()
    suppliers = Supplier.objects.prefetch_related('contacts').all()

    wb = openpyxl.Workbook()

    # ── Styles ────────────────────────────────────────────────────────────────
    hdr_fill  = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    hdr_font  = Font(bold=True, color='FFFFFF', size=11)
    sec_fill  = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
    sec_font  = Font(bold=True, color='FFFFFF', size=11)
    lbl_font  = Font(bold=True, size=10)
    title_font = Font(bold=True, color='FFFFFF', size=13)
    title_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    mfr_fill  = PatternFill(start_color='FFFBEA', end_color='FFFBEA', fill_type='solid')
    center    = Alignment(horizontal='center', vertical='center')
    left      = Alignment(horizontal='left',   vertical='center')

    # ── Sheet 1: Info ─────────────────────────────────────────────────────────
    ws_info = wb.active
    ws_info.title = 'Info'

    if not suppliers.exists():
        # Leave sheet empty with a placeholder note
        ws_info['A1'] = 'No supplier info available.'
        ws_info['A1'].font = Font(italic=True, color='888888')
    else:
        # Title row
        ws_info.merge_cells('A1:F1')
        t = ws_info['A1']
        t.value = 'ZEISS — Supplier Data Collection Form  |  Supplier Info & Contacts'
        t.font  = title_font
        t.fill  = title_fill
        t.alignment = center
        ws_info.row_dimensions[1].height = 22

        current_row = 2

        for sup_idx, supplier in enumerate(suppliers):
            contacts = list(supplier.contacts.all())

            # Blank separator between suppliers (skip before first)
            if sup_idx > 0:
                current_row += 1

            # ── Section A: Supplier General ───────────────────────────────
            ws_info.merge_cells(f'A{current_row}:F{current_row}')
            sec_a = ws_info[f'A{current_row}']
            sec_a.value = 'A  ·  SUPPLIER GENERAL'
            sec_a.font  = sec_font
            sec_a.fill  = sec_fill
            sec_a.alignment = left
            ws_info.row_dimensions[current_row].height = 18
            current_row += 1

            # Supplier Code row
            ws_info[f'A{current_row}'] = 'Supplier Code'
            ws_info[f'A{current_row}'].font = lbl_font
            ws_info.merge_cells(f'B{current_row}:F{current_row}')
            ws_info[f'B{current_row}'] = supplier.supplier_code
            current_row += 1

            # Company Name row
            ws_info[f'A{current_row}'] = 'Supplier Company Name:'
            ws_info[f'A{current_row}'].font = lbl_font
            ws_info.merge_cells(f'B{current_row}:F{current_row}')
            ws_info[f'B{current_row}'] = supplier.supplier_company_name
            current_row += 1

            current_row += 1  # blank row

            # ── Section B: Contact Information ────────────────────────────
            ws_info.merge_cells(f'A{current_row}:F{current_row}')
            sec_b = ws_info[f'A{current_row}']
            sec_b.value = 'B  ·  CONTACT INFORMATION  |  One row per person  ·  Repeat Contact Type for multiple contacts'
            sec_b.font  = sec_font
            sec_b.fill  = sec_fill
            sec_b.alignment = left
            ws_info.row_dimensions[current_row].height = 18
            current_row += 1

            # Contact table headers
            contact_headers = ['#', 'Contact Type', 'Name', 'Email', 'Phone', 'Role / Title']
            for ci, ch in enumerate(contact_headers, start=1):
                cell = ws_info.cell(row=current_row, column=ci, value=ch)
                cell.font = Font(bold=True, color='FFFFFF', size=10)
                cell.fill = hdr_fill
                cell.alignment = center
            ws_info.row_dimensions[current_row].height = 16
            current_row += 1

            # Contact rows — show all 6 types, filled or blank
            contact_map = {}
            for c in contacts:
                contact_map.setdefault(c.contact_type, []).append(c)

            row_num = 1
            for ct, _ in SupplierContact.CONTACT_TYPE_CHOICES:
                ct_contacts = contact_map.get(ct, [None])
                for c in ct_contacts:
                    ws_info.cell(row=current_row, column=1, value=row_num)
                    ws_info.cell(row=current_row, column=2, value=ct)
                    ws_info.cell(row=current_row, column=3, value=c.name       if c else '')
                    ws_info.cell(row=current_row, column=4, value=c.email      if c else '')
                    ws_info.cell(row=current_row, column=5, value=c.phone      if c else '')
                    ws_info.cell(row=current_row, column=6, value=c.role_title if c else '')
                    current_row += 1
                    row_num += 1

        # Column widths for Info sheet
        ws_info.column_dimensions['A'].width = 26
        ws_info.column_dimensions['B'].width = 32
        ws_info.column_dimensions['C'].width = 22
        ws_info.column_dimensions['D'].width = 30
        ws_info.column_dimensions['E'].width = 18
        ws_info.column_dimensions['F'].width = 22

    # ── Sheet 2: Materials ────────────────────────────────────────────────────
    ws_mat = wb.create_sheet(title='Materials')

    mat_headers = [
        'Supplier Code', 'Supplier Name', 'Part No', 'Part Description',
        'Order Qty', 'UOM', 'Unit Price', 'Currency',
        'PIC', 'Contact Email', 'Contact Secondary Email',
        'Lead Time (days)', 'Ship Lead Time (days)',
        'UOM (Quote)', 'COO', 'Currency (Quote)',
        'Unit Price 1', 'MOQ-1', 'Unit Price 2', 'MOQ-2', 'Unit Price 3', 'MOQ-3',
        'Lot Size', 'HTS Code', 'ECCN/EAR99',
        'Manufacture Part Number', 'Manufacturer Name',
        'Manufacturer Address', 'Item Weight (kg)', 'Volume Weight (kg)',
        'Russian Steel Confirmation', 'Hazmat (Y/N)', 'UN# · SDS/MSDS',
        'Product Regulation', 'EOL Status',
        'Alternative Part(s)', 'Alternative Part No',
        'Mfg Address + Postal (CN Only)', 'UFLPA Compliance (CN Only)',
        'UFLPA Start Date', 'UFLPA Expiry Date',
        'USMCA Certificate (CA/MX Only)', 'USMCA Start Date', 'USMCA Expiry Date',
        'Status', 'Comments',
    ]

    for col_idx, header in enumerate(mat_headers, start=1):
        cell = ws_mat.cell(row=1, column=col_idx, value=header)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center

    for row_idx, entry in enumerate(entries, start=2):
        row = [
            entry.supplier_code,
            entry.supplier_name,
            entry.part_no,
            entry.part_description,
            float(entry.order_qty)       if entry.order_qty       is not None else '',
            entry.uom,
            float(entry.unit_price)      if entry.unit_price      is not None else '',
            entry.currency,
            entry.pic,
            entry.contact_email,
            entry.contact_secondary_email,
            entry.lead_time_days         or '',
            entry.ship_lead_time_days    or '',
            entry.quote_uom,
            entry.coo,
            entry.quote_currency,
            float(entry.unit_price_1)    if entry.unit_price_1    is not None else '',
            float(entry.moq_1)           if entry.moq_1           is not None else '',
            float(entry.unit_price_2)    if entry.unit_price_2    is not None else '',
            float(entry.moq_2)           if entry.moq_2           is not None else '',
            float(entry.unit_price_3)    if entry.unit_price_3    is not None else '',
            float(entry.moq_3)           if entry.moq_3           is not None else '',
            float(entry.lot_size)        if entry.lot_size        is not None else '',
            entry.hts_code,
            entry.eccn_ear99,
            entry.manufacture_part_number,
            entry.manufacturer_name,
            entry.manufacturer_address,
            float(entry.item_weight_kg)  if entry.item_weight_kg  is not None else '',
            float(entry.volume_weight_kg)if entry.volume_weight_kg is not None else '',
            entry.russian_steel_confirmation,
            entry.hazmat,
            entry.un_sds_msds,
            entry.product_regulation,
            entry.eol_status,
            entry.alternative_parts,
            entry.alternative_part_no,
            entry.mfg_address_postal_cn,
            entry.uflpa_compliance,
            entry.uflpa_start_date,
            entry.uflpa_expiry_date,
            entry.usmca_certificate,
            entry.usmca_start_date,
            entry.usmca_expiry_date,
            entry.status,
            entry.comments,
        ]
        for col_idx, value in enumerate(row, start=1):
            cell = ws_mat.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 26:   # Manufacture Part Number
                cell.fill = mfr_fill

    # Auto-fit Materials columns
    for col_idx, header in enumerate(mat_headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for row_idx in range(2, ws_mat.max_row + 1):
            val = ws_mat.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws_mat.column_dimensions[col_letter].width = min(max_len + 4, 50)

    ws_mat.freeze_panes = 'A2'
    ws_mat.row_dimensions[1].height = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="RFQ_Tracker_Export.xlsx"'
    wb.save(response)
    return response


def _parse_info_sheet(ws):
    """
    Parse the Info sheet from a ZEISS Supplier Template workbook.
    Returns (supplier_code, company_name, contacts_list) or raises ValueError.
    """
    def _n(val):
        if val is None:
            return ''
        t = str(val)
        for ch in '▼▾▽▶↓▴':
            t = t.replace(ch, '')
        return ' '.join(t.split()).strip()

    supplier_code = ''
    company_name = ''
    contact_header_row = None
    col_map = {}
    all_rows = list(ws.iter_rows(values_only=True))

    for row_idx, row_vals in enumerate(all_rows, start=1):
        for col_idx, raw in enumerate(row_vals):
            nval = _n(raw)
            nl = nval.lower()

            if nl in ('supplier code', 'supplier code:') and not supplier_code:
                for v2 in row_vals[col_idx + 1:]:
                    nv2 = _n(v2)
                    if nv2:
                        supplier_code = nv2
                        break

            if 'supplier company name' in nl and not company_name:
                for v2 in row_vals[col_idx + 1:]:
                    nv2 = _n(v2)
                    if nv2:
                        company_name = nv2
                        break

            if nl == 'contact type' and contact_header_row is None:
                contact_header_row = row_idx
                for ci, hraw in enumerate(row_vals):
                    hv = _n(hraw).lower()
                    if hv == 'contact type':
                        col_map['contact_type'] = ci
                    elif hv == 'name':
                        col_map['name'] = ci
                    elif hv == 'email':
                        col_map['email'] = ci
                    elif hv == 'phone':
                        col_map['phone'] = ci
                    elif 'role' in hv or 'title' in hv:
                        col_map['role_title'] = ci

    contacts = []
    VALID_TYPES = {ct for ct, _ in SupplierContact.CONTACT_TYPE_CHOICES}
    ct_col = col_map.get('contact_type')
    if contact_header_row is not None and ct_col is not None:
        for row_vals in all_rows[contact_header_row:]:
            ct_val = _n(row_vals[ct_col]) if ct_col < len(row_vals) else ''
            if not ct_val or ct_val.isdigit():
                continue
            matched_type = ct_val
            for vt in VALID_TYPES:
                if ct_val.lower() in vt.lower() or vt.lower() in ct_val.lower():
                    matched_type = vt
                    break

            def _gv(key):
                ci = col_map.get(key)
                return _n(row_vals[ci]) if ci is not None and ci < len(row_vals) else ''

            contacts.append({
                'contact_type': matched_type,
                'name':         _gv('name'),
                'email':        _gv('email'),
                'phone':        _gv('phone'),
                'role_title':   _gv('role_title'),
            })

    return supplier_code, company_name, contacts


def _entries_identical(existing, incoming_kwargs):
    """Return True when every incoming field matches the existing DB row.
    Status and comments are skipped — they are manually maintained.

    Uses the same serialisation path as _entry_to_dict / _kwargs_to_display
    so that Decimal('6.00') and float 6.0 are treated as equal.
    """
    SKIP = {'status', 'comments'}

    def _norm(field, s):
        """Normalise a serialised string value for comparison."""
        s = (s or '').strip()
        if field in DECIMAL_FIELDS or field in INT_FIELDS:
            try:
                return str(float(s))   # '6.00'→'6.0', '51.0500'→'51.05'
            except (ValueError, TypeError):
                pass
        return s

    existing_dict  = _entry_to_dict(existing)        # Decimal/date → string
    incoming_strs  = _kwargs_to_display(incoming_kwargs)  # float/date → string

    for field, new_str in incoming_strs.items():
        if field in SKIP:
            continue
        existing_str = existing_dict.get(field, '')
        if _norm(field, existing_str) != _norm(field, new_str):
            return False
    return True


def _kwargs_to_display(kwargs):
    """Serialize a coerced kwargs dict to JSON-safe strings (mirrors _entry_to_dict format)."""
    def _s(v):
        if v is None:
            return ''
        if isinstance(v, datetime.datetime):
            return v.date().isoformat()   # "2021-01-29", not "2021-01-29T00:00:00"
        if isinstance(v, datetime.date):
            return v.isoformat()
        return str(v)
    return {k: _s(v) for k, v in kwargs.items()}


@login_required
def rfq_resolve_duplicates(request):
    """Apply keep/replace choices returned by the duplicate resolution modal."""
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'ok': False, 'error': 'Invalid JSON'}, status=400)

    replaced = 0
    for res in data.get('resolutions', []):
        if res.get('action') != 'replace':
            continue
        pk = res.get('pk')
        if not pk:
            continue
        try:
            entry = RFQEntry.objects.get(pk=int(pk))
        except (RFQEntry.DoesNotExist, ValueError):
            continue
        for field, raw in res.get('data', {}).items():
            if hasattr(entry, field) and field not in ('pk', 'id'):
                setattr(entry, field, _coerce(field, raw if raw != '' else None))
        entry.save()
        replaced += 1

    return JsonResponse({'ok': True, 'replaced': replaced})


@login_required
def rfq_bulk_upload(request):
    if request.method != 'POST':
        return redirect('rfq_list')

    files = request.FILES.getlist('excel_files')
    if not files:
        messages.error(request, 'No files selected.')
        return redirect('rfq_list')

    total_added = 0
    sup_created = 0
    sup_updated = 0
    errors = []
    duplicates = []
    identical_skipped = 0

    for f in files:
        try:
            wb = openpyxl.load_workbook(f, data_only=True)

            # ── Sheet 1 (Info): parse supplier info if present ────────────
            if 'Info' in wb.sheetnames:
                ws_info = wb['Info']
                try:
                    supplier_code, company_name, contacts = _parse_info_sheet(ws_info)
                    if supplier_code:
                        supplier, was_created = Supplier.objects.get_or_create(
                            supplier_code=supplier_code,
                            defaults={'supplier_company_name': company_name},
                        )
                        if not was_created:
                            supplier.supplier_company_name = company_name
                            supplier.save()
                            sup_updated += 1
                        else:
                            sup_created += 1

                        supplier.contacts.all().delete()
                        for c in contacts:
                            if not any([c['name'], c['email'], c['phone'], c['role_title']]):
                                continue
                            SupplierContact.objects.create(
                                supplier=supplier,
                                contact_type=c['contact_type'],
                                name=c['name'],
                                email=c['email'],
                                phone=c['phone'],
                                role_title=c['role_title'],
                            )
                except Exception as e:
                    errors.append(f"{f.name} (Info sheet): {e}")

            # ── Sheet 2 (Materials): parse RFQ entries ────────────────────
            if 'Materials' in wb.sheetnames:
                ws = wb['Materials']
            else:
                ws = wb.active

            # ── Build header → column-index map ──────────────────────
            headers = {}
            row1_cells = list(ws[1])
            row2_cells = list(ws[2]) if ws.max_row >= 2 else []
            data_start_row = 2
            uses_row2_header = False

            for ci, cell in enumerate(row1_cells):
                r1 = _normalize_header(cell.value)
                field = _resolve_header(r1)

                if not field and ci < len(row2_cells) and row2_cells[ci].value:
                    r2 = _normalize_header(row2_cells[ci].value)
                    combined = f"{r1} {r2}".strip()
                    field = _resolve_header(combined)
                    if field:
                        uses_row2_header = True

                if field:
                    if field in headers:
                        field = DUPLICATE_FIELD_MAP.get(field)
                    if field and field not in headers:
                        headers[field] = ci + 1

            if uses_row2_header:
                data_start_row = 3

            if 'supplier_code' not in headers and 'supplier_name' not in headers:
                errors.append(f"{f.name}: could not find recognisable column headers in Materials sheet.")
                continue

            added = 0
            for row in ws.iter_rows(min_row=data_start_row, values_only=True):
                if all(v is None or str(v).strip() == '' for v in row):
                    continue

                kwargs = {}
                for field, col_idx in headers.items():
                    raw = row[col_idx - 1] if col_idx - 1 < len(row) else None
                    kwargs[field] = _coerce(field, raw)

                if not kwargs.get('supplier_code') and not kwargs.get('supplier_name'):
                    continue

                existing = RFQEntry.objects.filter(
                    supplier_code=kwargs.get('supplier_code') or '',
                    supplier_name=kwargs.get('supplier_name') or '',
                    part_no=kwargs.get('part_no') or '',
                ).first()

                if existing:
                    if _entries_identical(existing, kwargs):
                        identical_skipped += 1   # no real change — skip silently
                    else:
                        duplicates.append({
                            'existing': _entry_to_dict(existing),
                            'incoming': _kwargs_to_display(kwargs),
                        })
                else:
                    RFQEntry.objects.create(**kwargs)
                    added += 1

            total_added += added

        except Exception as e:
            errors.append(f"{f.name}: {e}")

    # AJAX request — return JSON so the frontend can show duplicate resolution UI
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'ok': True,
            'added': total_added,
            'identical_skipped': identical_skipped,
            'duplicates': duplicates,
            'errors': errors,
        })

    # Fallback plain POST (no AJAX) — original redirect behaviour
    parts = []
    if total_added:
        parts.append(f'{total_added} RFQ row(s) added')
    if sup_created:
        parts.append(f'{sup_created} supplier(s) created')
    if sup_updated:
        parts.append(f'{sup_updated} supplier(s) updated')
    if parts:
        messages.success(request, f'Upload complete — {", ".join(parts)}.')
    if errors:
        for err in errors:
            messages.error(request, err)
    return redirect('rfq_list')


@login_required
def rfq_deduplicate(request):
    """
    GET  → return count of duplicate rows that would be removed (dry-run).
    POST → remove them, keeping the entry with the most filled fields per group.
    """
    from django.db.models import Count

    groups = (
        RFQEntry.objects
        .values('supplier_code', 'supplier_name', 'part_no')
        .annotate(cnt=Count('pk'))
        .filter(cnt__gt=1)
    )

    SCORE_FIELDS = ['status', 'comments', 'contact_email', 'pic',
                    'unit_price', 'lead_time_days', 'coo']

    def _score(entry):
        return sum(1 for f in SCORE_FIELDS if getattr(entry, f, None) not in (None, ''))

    if request.method == 'GET':
        total_dupes = sum(g['cnt'] - 1 for g in groups)
        return JsonResponse({'ok': True, 'duplicate_rows': total_dupes})

    removed = 0
    for g in groups:
        entries = list(RFQEntry.objects.filter(
            supplier_code=g['supplier_code'],
            supplier_name=g['supplier_name'],
            part_no=g['part_no'],
        ).order_by('pk'))
        best = max(entries, key=lambda e: (_score(e), -e.pk))
        pks_to_delete = [e.pk for e in entries if e.pk != best.pk]
        removed += RFQEntry.objects.filter(pk__in=pks_to_delete).delete()[0]

    return JsonResponse({'ok': True, 'removed': removed})


@login_required
def rfq_clear_all(request):
    """Delete every RFQEntry (POST only)."""
    if request.method == 'POST':
        count, _ = RFQEntry.objects.all().delete()
        messages.success(request, f'All data cleared — {count} record(s) deleted.')
    return redirect('rfq_list')


@login_required
def rfq_bulk_status(request):
    """Update status on a list of RFQ entries via AJAX JSON POST."""
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    try:
        data   = json.loads(request.body)
        pks    = [int(p) for p in data.get('pks', [])]
        status = (data.get('status') or '').strip()
    except (json.JSONDecodeError, ValueError, AttributeError):
        return JsonResponse({'ok': False, 'error': 'Invalid request.'}, status=400)
    if not pks:
        return JsonResponse({'ok': False, 'error': 'No rows selected.'}, status=400)
    updated = RFQEntry.objects.filter(pk__in=pks).update(status=status)
    return JsonResponse({'ok': True, 'updated': updated})


@login_required
def rfq_patch(request, pk):
    """Inline-update a single dropdown field via AJAX."""
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    entry = get_object_or_404(RFQEntry, pk=pk)
    try:
        data = json.loads(request.body)
        field = data.get('field', '')
        value = data.get('value', '')
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'ok': False, 'error': 'Invalid JSON'}, status=400)
    PATCHABLE = {
        'russian_steel_confirmation', 'hazmat', 'un_sds_msds',
        'product_regulation', 'eol_status', 'uflpa_compliance', 'usmca_certificate',
        'status', 'comments',
    }
    if field not in PATCHABLE:
        return JsonResponse({'ok': False, 'error': 'Field not allowed'}, status=400)
    setattr(entry, field, value)
    entry.save(update_fields=[field])
    return JsonResponse({'ok': True})


@login_required
def rfq_edit_json(request, pk):
    """Load (GET) or save (POST) an entry via AJAX — used by the edit modal."""
    entry = get_object_or_404(RFQEntry, pk=pk)
    if request.method == 'POST':
        form = RFQEntryForm(request.POST, instance=entry, prefix='edit')
        if form.is_valid():
            saved = form.save()
            return JsonResponse({'ok': True, 'entry': _entry_to_dict(saved)})
        errors = {k: [str(e) for e in v] for k, v in form.errors.items()}
        return JsonResponse({'ok': False, 'errors': errors}, status=400)
    return JsonResponse({'ok': True, 'entry': _entry_to_dict(entry)})


# ── Supplier Info Views ────────────────────────────────────────────────────────

def _supplier_to_dict(supplier):
    contacts = list(supplier.contacts.values(
        'id', 'contact_type', 'name', 'email', 'phone', 'role_title'
    ))
    return {
        'pk': supplier.pk,
        'supplier_code': supplier.supplier_code,
        'supplier_company_name': supplier.supplier_company_name,
        'contact_count': len(contacts),
        'contacts': contacts,
    }


@login_required
def supplier_list(request):
    total = Supplier.objects.count()
    return render(request, 'tracker/supplier_list.html', {'total': total})


@login_required
def supplier_data(request):
    from django.db.models import Q
    pk = request.GET.get('pk')
    if pk:
        try:
            s = Supplier.objects.get(pk=int(pk))
            return JsonResponse({'suppliers': [_supplier_to_dict(s)], 'total': 1})
        except (Supplier.DoesNotExist, ValueError):
            return JsonResponse({'suppliers': [], 'total': 0})
    q = request.GET.get('q', '').strip()
    qs = Supplier.objects.all()
    if q:
        qs = qs.filter(
            Q(supplier_code__icontains=q) | Q(supplier_company_name__icontains=q)
        )
    suppliers = [_supplier_to_dict(s) for s in qs]
    return JsonResponse({'suppliers': suppliers, 'total': len(suppliers)})


@login_required
def supplier_save(request):
    """Create or update a supplier + its contacts via AJAX JSON POST."""
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'ok': False, 'error': 'Invalid JSON'}, status=400)

    pk = data.get('pk')
    supplier_code = (data.get('supplier_code') or '').strip()
    supplier_company_name = (data.get('supplier_company_name') or '').strip()
    contacts_data = data.get('contacts', [])

    if not supplier_code:
        return JsonResponse({'ok': False, 'error': 'Supplier Code is required.'}, status=400)

    if pk:
        supplier = get_object_or_404(Supplier, pk=pk)
        # Check uniqueness if code changed
        if Supplier.objects.filter(supplier_code=supplier_code).exclude(pk=pk).exists():
            return JsonResponse({'ok': False, 'error': f'Supplier Code "{supplier_code}" already exists.'}, status=400)
        supplier.supplier_code = supplier_code
        supplier.supplier_company_name = supplier_company_name
        supplier.save()
        supplier.contacts.all().delete()
    else:
        if Supplier.objects.filter(supplier_code=supplier_code).exists():
            return JsonResponse({'ok': False, 'error': f'Supplier Code "{supplier_code}" already exists.'}, status=400)
        supplier = Supplier.objects.create(
            supplier_code=supplier_code,
            supplier_company_name=supplier_company_name,
        )

    VALID_TYPES = {ct for ct, _ in SupplierContact.CONTACT_TYPE_CHOICES}
    for c in contacts_data:
        ct = (c.get('contact_type') or '').strip()
        name = (c.get('name') or '').strip()
        email = (c.get('email') or '').strip()
        phone = (c.get('phone') or '').strip()
        role_title = (c.get('role_title') or '').strip()
        if not ct or ct not in VALID_TYPES:
            continue
        if not any([name, email, phone, role_title]):
            continue
        SupplierContact.objects.create(
            supplier=supplier,
            contact_type=ct,
            name=name,
            email=email,
            phone=phone,
            role_title=role_title,
        )

    return JsonResponse({'ok': True, 'supplier': _supplier_to_dict(supplier)})


@login_required
def supplier_delete(request, pk):
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    supplier = get_object_or_404(Supplier, pk=pk)
    supplier.delete()
    return JsonResponse({'ok': True})


@login_required
def supplier_template_upload(request):
    """Parse Info sheet(s) from uploaded ZEISS Supplier Template files."""
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)

    files = request.FILES.getlist('excel_files')
    if not files:
        return JsonResponse({'ok': False, 'error': 'No files provided.'})

    created, updated, errors = 0, 0, []

    for f in files:
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb['Info'] if 'Info' in wb.sheetnames else wb.worksheets[0]

            supplier_code, company_name, contacts = _parse_info_sheet(ws)

            if not supplier_code:
                errors.append(f"{f.name}: Could not find 'Supplier Code' in the Info sheet.")
                continue

            supplier, was_created = Supplier.objects.get_or_create(
                supplier_code=supplier_code,
                defaults={'supplier_company_name': company_name},
            )
            if not was_created:
                supplier.supplier_company_name = company_name
                supplier.save()
                updated += 1
            else:
                created += 1

            supplier.contacts.all().delete()
            for c in contacts:
                if not any([c['name'], c['email'], c['phone'], c['role_title']]):
                    continue
                SupplierContact.objects.create(
                    supplier=supplier,
                    contact_type=c['contact_type'],
                    name=c['name'],
                    email=c['email'],
                    phone=c['phone'],
                    role_title=c['role_title'],
                )

        except Exception as e:
            errors.append(f"{f.name}: {e}")

    parts = []
    if created: parts.append(f"{created} supplier(s) created")
    if updated: parts.append(f"{updated} supplier(s) updated")
    msg = ', '.join(parts) + '.' if parts else 'No suppliers imported.'

    return JsonResponse({'ok': True, 'message': msg, 'errors': errors})
